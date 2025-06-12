import pandas as pd
import numpy as np
from pathlib import Path
import time
from datetime import datetime
from typing import Optional, Dict, List, Tuple
import logging
import re
import unicodedata
from openpyxl import load_workbook
from address_parser import OptimizedAddressParser

class ExcelAddressProcessor:
    """
    Procesador optimizado para archivos Excel con direcciones - PRESERVA FORMATO
    """
    
    def __init__(self, log_level: str = 'INFO'):
        self.parser = OptimizedAddressParser()
        self.setup_logging(log_level)
        self.processed_count = 0
        self.error_count = 0
        self.stats = {
            'total_rows': 0,
            'processed': 0,
            'errors': 0,
            'empty_addresses': 0,
            'processing_time': 0
        }
    
    def setup_logging(self, level: str):
        """Configurar logging solo en consola, sin archivo .log"""
        log_level = getattr(logging, level.upper(), logging.INFO)
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def validate_file(self, file_path: str) -> Tuple[bool, str]:
        """Validar archivo"""
        path = Path(file_path)
        
        if not path.exists():
            return False, f"El archivo {file_path} no existe"
        
        if not path.suffix.lower() in ['.xlsx', '.xls']:
            return False, f"El archivo debe ser Excel (.xlsx o .xls)"
        
        try:
            pd.read_excel(file_path, nrows=1, engine='openpyxl')
            return True, "Archivo válido"
        except Exception as e:
            return False, f"Error leyendo el archivo: {str(e)}"
    
    def normalize_text(self, text: str) -> str:
        """Normaliza el texto: quita acentos y convierte a mayúsculas"""
        if not isinstance(text, str):
            text = str(text)
        text = text.upper().strip()
        text = ''.join(c for c in unicodedata.normalize('NFD', text)
                      if unicodedata.category(c) != 'Mn')
        return text

    def get_column_letter(self, col_index: int) -> str:
        """Convierte índice de columna (0-based) a letra de columna de Excel"""
        result = ""
        while col_index >= 0:
            result = chr(col_index % 26 + ord('A')) + result
            col_index = col_index // 26 - 1
        return result

    def detect_address_column(self, df: pd.DataFrame) -> Optional[str]:
        """Detectar columna de direcciones"""
        target_names = ['DIRECCION PRINCIPAL', 'DIRECCION PRINCI']
        
        # Imprimir todas las columnas para debug
        print("\nColumnas encontradas:")
        for idx, col in enumerate(df.columns):
            normalized = self.normalize_text(str(col))
            col_letter = self.get_column_letter(idx)
            print(f"{col_letter}: {col} -> {normalized}")
            
        # Buscar coincidencia por nombre normalizado
        for idx, col in enumerate(df.columns):
            normalized = self.normalize_text(str(col))
            if any(target in normalized for target in target_names):
                col_letter = self.get_column_letter(idx)
                self.logger.info(f"Columna de direcciones encontrada: {col} en columna {col_letter}")
                return col
        
        return None

    def detect_param_column(self, df: pd.DataFrame) -> Tuple[Optional[str], Optional[int]]:
        """Detectar columna de parametrización"""
        target_names = ['PARAMETRIZACION', 'PRAMETRIZACION']
        
        # Buscar coincidencia por nombre normalizado
        for idx, col in enumerate(df.columns):
            normalized = self.normalize_text(str(col))
            if any(target in normalized for target in target_names):
                col_letter = self.get_column_letter(idx)
                self.logger.info(f"Columna de parametrización encontrada: {col} en columna {col_letter}")
                return col, idx
                
        return None, None

    def normalize_colname(self, colname: str) -> str:
        """Normaliza el nombre de columna quitando tildes y pasando a mayúsculas"""
        return ''.join(c for c in unicodedata.normalize('NFD', colname)
                       if unicodedata.category(c) != 'Mn').upper()

    def process_excel_file_preserve_format(self, file_path: str = "Backlog_GPON_FILTRADO.xlsx", show_progress: bool = True) -> bool:
        try:
            # Validar archivo
            is_valid, message = self.validate_file(file_path)
            if not is_valid:
                self.logger.error(message)
                return False
            
            print(f"🔄 Procesando archivo: {file_path}")
            print("⚠️  MODO PRESERVACIÓN DE FORMATO - Solo se editará el contenido de celdas")
            
            # Usar openpyxl para preservar formato
            workbook = load_workbook(file_path)
            if "GPON" not in workbook.sheetnames:
                self.logger.error("La hoja 'GPON' no existe en el archivo")
                return False
            worksheet = workbook["GPON"]
            
            # También leer con pandas para facilitar la detección de columnas
            df = pd.read_excel(file_path, sheet_name="GPON", engine='openpyxl')
            
            # NO normalizar nombres de columnas aquí, trabajar con los originales
            print(f"Total de columnas en la hoja 'GPON': {len(df.columns)}")
            
            # Detectar columna de direcciones
            address_col_name = self.detect_address_column(df)
            if not address_col_name:
                self.logger.error("No se pudo detectar la columna de direcciones")
                return False
            
            # Detectar columna de parametrización
            param_col_name, param_col_idx = self.detect_param_column(df)
            
            # Si no se encuentra, mostrar opciones y pedir al usuario
            if not param_col_name:
                print("\n❌ No se encontró columna de parametrización automáticamente.")
                print("Columnas disponibles:")
                for i, col in enumerate(df.columns):
                    col_letter = self.get_column_letter(i)
                    print(f"  {col_letter}: {col}")
                print("Por favor, ingrese la letra de la columna de parametrización (ejemplo: Z): ", end='')
                col_letter = input().strip().upper()
                
                # Convertir letra a índice (solo para columnas A-Z y AA-AZ)
                if len(col_letter) == 1:
                    param_col_idx = ord(col_letter) - ord('A')
                elif len(col_letter) == 2 and col_letter[0] == 'A':
                    param_col_idx = 26 + ord(col_letter[1]) - ord('A')
                else:
                    self.logger.error("Solo se soportan columnas A-Z y AA-AZ")
                    return False
                
                if 0 <= param_col_idx < len(df.columns):
                    param_col_name = df.columns[param_col_idx]
                    print(f"Columna seleccionada: {param_col_name}")
                else:
                    self.logger.error("Letra de columna inválida.")
                    return False
        
            # Encontrar índice de la columna de direcciones
            address_col_idx = df.columns.get_loc(address_col_name)
            
            address_letter = self.get_column_letter(address_col_idx)
            param_letter = self.get_column_letter(param_col_idx)
            
            print(f"📍 Columna de direcciones: {address_col_name} (columna {address_letter})")
            print(f"🎯 Columna de parametrización: {param_col_name} (columna {param_letter})")
            
            # Verificar que los índices sean correctos
            print(f"🔍 Verificación:")
            print(f"   - Índice de direcciones: {address_col_idx} (debería ser {self.get_column_letter(address_col_idx)})")
            print(f"   - Índice de parametrización: {param_col_idx} (debería ser {self.get_column_letter(param_col_idx)})")
            
            # Procesar solo filas visibles
            print("🔄 Procesando direcciones solo en filas visibles...")
            modificadas = 0
            vaciadas = 0
            errores = 0
            
            for row_idx in range(2, worksheet.max_row + 1):
                # Verificar si la fila está oculta
                if worksheet.row_dimensions[row_idx].hidden:
                    continue
                
                # Obtener la celda de dirección (índice + 1 porque openpyxl es 1-based)
                direccion_cell = worksheet.cell(row=row_idx, column=address_col_idx + 1)
                direccion_original = str(direccion_cell.value or "").strip()
                
                if not direccion_original or direccion_original.lower() in ['none', 'nan']:
                    continue
                
                try:
                    # CORECCIÓN PRINCIPAL: Usar parametrizar_direccion() en lugar de parse_address()
                    result = self.parser.parametrizar_direccion(direccion_original)
                    
                    # Obtener la celda de parametrización
                    param_cell = worksheet.cell(row=row_idx, column=param_col_idx + 1)
                    
                    if result['valida'] and result['direccion_parametrizada'] != "DIRECCION NO VALIDA":
                        param_cell.value = result['direccion_parametrizada']
                        modificadas += 1
                        
                        # Debug: mostrar algunas conversiones
                        if modificadas <= 5:
                            print(f"   ✅ Fila {row_idx}: '{direccion_original}' -> '{result['direccion_parametrizada']}'")
                    else:
                        # Si la dirección no es válida, limpiar la celda
                        if param_cell.value is not None:
                            param_cell.value = None
                            vaciadas += 1
                        
                        # Debug: mostrar direcciones problemáticas
                        if errores < 3:
                            print(f"   ❌ Fila {row_idx}: '{direccion_original}' -> No válida ({'; '.join(result['errores'])})")
                        errores += 1
                        
                except Exception as e:
                    errores += 1
                    if errores <= 3:
                        print(f"   ⚠️  Error fila {row_idx}: {str(e)}")
            
            # Guardar el archivo SIN cambiar formato
            print("💾 Guardando cambios...")
            workbook.save(file_path)
            workbook.close()
            
            # Resumen final
            print(f"\n{'='*60}")
            print("✅ PROCESAMIENTO COMPLETADO")
            print(f"{'='*60}")
            print(f"📁 Archivo: {file_path}")
            print(f"✏️  Celdas modificadas: {modificadas}")
            print(f"🧹 Celdas vaciadas: {vaciadas}")
            print(f"❌ Errores encontrados: {errores}")
            print(f"🎨 Formato original preservado: ✅")
            print(f"{'='*60}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error procesando archivo: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def process_excel_file(self, file_path: str = "Backlog250525-filtradocobertura.xlsx", 
                          show_progress: bool = True) -> bool:
        """
        Método alias para mantener compatibilidad - llama al método principal
        """
        return self.process_excel_file_preserve_format(file_path, show_progress)

    def get_addresses_from_column(self, file_path, column, sheet_name='GPON'):
        """Obtener direcciones de una columna específica"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            return df[column].dropna().tolist()
        except KeyError:
            print(f"\n⚠️  Columna '{column}' no encontrada en la hoja {sheet_name}")
            print("📊 Columnas disponibles:")
            for col in df.columns:
                print(f"   - {col}")
            raise
        except Exception as e:
            print(f"❌ Error leyendo el archivo: {str(e)}")
            return []

    def get_column_names(self, excel_file: str) -> List[str]:
        """Retorna lista de nombres de columnas disponibles en el archivo"""
        try:
            df = pd.read_excel(excel_file)
            return df.columns.tolist()
        except Exception as e:
            print(f"Error obteniendo nombres de columnas: {str(e)}")
            return []


# Código para ejecutar el procesamiento
if __name__ == "__main__":
    processor = ExcelAddressProcessor()
    
    # Procesar el archivo PRESERVANDO FORMATO
    success = processor.process_excel_file("Backlog_GPON_FILTRADO.xlsx")
    
    if success:
        print("\n✅ Procesamiento completado exitosamente")
        print("🎨 El formato original se ha preservado completamente")
    else:
        print("\n❌ Error en el procesamiento")