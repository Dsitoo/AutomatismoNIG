from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import os
import glob
import shutil
import time
import pandas as pd
from datetime import datetime, timedelta
import xlrd  # Agregar esta importación
import openpyxl # Agregar esta importación

class AddressAutomation:
    def __init__(self):
        self.firefox_path = r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe"
        self.url = "http://10.112.1.55:8080/NIG-Client-PB/"
        self.username = "clauhert"
        self.password = "clauhert"
        self.download_folder = os.path.expanduser('~/Downloads')
        self.results_folder = 'C:/Resultados'
        self.addresses = self.load_addresses_from_excel()
        self.current_address = None
        # Agregar diccionario para almacenar información
        self.address_data = {}  # Estructura: {dirección: {datos...}}
        self.not_found_addresses = []  # Lista de direcciones no encontradas
        self.found_addresses = {}  # Diccionario para almacenar {dirección_encontrada: dirección_original}
        # Eliminar default_address ya que usaremos current_address
        self.setup_driver()

    def load_addresses_from_excel(self):
        """Cargar direcciones desde el archivo Excel"""
        try:
            excel_file = "Backlog_GPON_FILTRADO.xlsx"
            df = pd.read_excel(excel_file, sheet_name='GPON')
            
            # Definir las columnas requeridas
            required_columns = [
                'Dentro molecula Gpon',
                'Asociado molecula',
                'Propietario Molecula',
                'Nombre de la moleula',
                'Estado Molecula'
            ]
            
            def needs_processing(row):
                """Verificar si una fila necesita ser procesada"""
                values = [str(row[col]).strip().upper() for col in required_columns]
                
                # Si todos los campos están vacíos o son N/A, necesita procesamiento
                all_empty_or_na = all(v in ['', 'N/A', 'NAN', 'NONE', 'NULL'] for v in values)
                if all_empty_or_na:
                    return True
                
                # Si al menos un campo está vacío o es N/A, necesita procesamiento
                any_empty_or_na = any(v in ['', 'N/A', 'NAN', 'NONE', 'NULL'] for v in values)
                if any_empty_or_na:
                    return True
                
                # Caso especial: Si 'Dentro molecula Gpon' es 'No', verificar que los demás campos sean 'N/A'
                # excepto 'Asociado molecula' que debe ser 'No'
                if values[0] == 'NO':  # Dentro molecula Gpon
                    expected = ['NO', 'NO', 'N/A', 'N/A', 'N/A']
                    return values != expected
                
                # Si todos los campos tienen información válida, no necesita procesamiento
                return False
            
            # Filtrar direcciones que necesitan procesamiento
            mask = df.apply(needs_processing, axis=1)
            unique_addresses = df.loc[mask, 'Prametrización'].drop_duplicates().dropna().tolist()
            
            # Crear diccionario de índices por dirección para procesamiento posterior
            self.address_indices = {}
            for idx, row in df.iterrows():
                addr = row['Prametrización']
                if addr in unique_addresses:
                    if addr not in self.address_indices:
                        self.address_indices[addr] = []
                    self.address_indices[addr].append(idx)
            
            print(f"\n📊 Análisis de direcciones:")
            print(f"✅ Total de direcciones únicas por procesar: {len(unique_addresses)}")
            print(f"ℹ️ Direcciones omitidas por tener información completa: {len(df) - len(unique_addresses)}")
            
            if len(unique_addresses) == 0:
                print("ℹ️ No se encontraron direcciones que requieran procesamiento")
            
            return unique_addresses
            
        except Exception as e:
            print(f"❌ Error cargando direcciones: {str(e)}")
            return []

    def login(self):
        """Realizar el proceso de login"""
        try:
            # Esperar a que la tabla de login esté presente
            WebDriverWait(self.driver, 10).until(
                lambda d: d.find_element_by_class_name("login_table")
            )
            
            # Login
            print("Iniciando sesión...")
            self.driver.find_element_by_name("j_username").send_keys(self.username)
            self.driver.find_element_by_name("j_password").send_keys(self.password)
            self.driver.find_element_by_name("login").click()
            
            print("Login exitoso")
            time.sleep(1)  # Esperar que cargue post-login
        except Exception as e:
            print(f"Error en login: {str(e)}")
            raise

    def handle_silverlight(self):
        """Manejar activación de Silverlight"""
        try:
            
            time.sleep(2)
            
            # Obtener la posición y tamaño de la ventana activa del navegador
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Calcular el centro de la ventana
            center_x = window['x'] + (size['width'] // 2)
            center_y = window['y'] + (size['height'] // 2)
            
            # Primer clic en el centro del navegador
            print("Activando permisos de SilverLight...")
            pyautogui.click(center_x, center_y)
            time.sleep(1.5)
            
            # Clic en el botón "Allow Now" - está en la parte superior izquierda
            
            # El botón aparece aproximadamente a 115px desde la izquierda y 90px desde arriba
            allow_x = window['x'] + 110
            allow_y = window['y'] + 150
            pyautogui.click(allow_x, allow_y)
            
            print("Silverlight activado")
            
            # Preguntar si la interfaz cargó correctamente después de activar Silverlight
            while True:
                response = input("\n¿La interfaz cargó correctamente? (s/n): ").lower()
                if response == 'n':
                    print("Deteniendo automatización por fallo en carga de interfaz...")
                    self.driver.quit()
                    exit()
                elif response == 's':
                    print("Continuando automatización en 5 segundos...")
                    time.sleep(5)
                    break
                else:
                    print("Por favor responda 's' para sí o 'n' para no")
                    
        except Exception as e:
            print(f"Error activando Silverlight: {str(e)}")

    def click_default_button(self):
        """Hacer clic en el botón de la barra de herramientas"""
        try:
            
            time.sleep(2)
            
            # Obtener posición de la ventana
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Coordenadas ajustadas - botón está a aproximadamente 85% del ancho total
            # y 35px desde arriba en la barra de herramientas
            button_x = window['x'] + int(size['width'] * 0.84)  # 85% del ancho
            button_y = window['y'] + 120  # Altura de la barra de herramientas
            
           
            pyautogui.click(button_x, button_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en botón: {str(e)}")

    def click_new_query(self):
        """Hacer clic en el botón New Query usando coordenadas"""
        try:
            print("Iniciando formulario de busqueda...")
            time.sleep(3)  # Dar más tiempo para que cargue
            
            # Obtener posición de la ventana
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # El botón "New Query" está aproximadamente en la parte central-superior
            query_x = window['x'] + int(size['width'] * 0.45)  # 50% del ancho
            query_y = window['y'] + 160  # Misma altura que el botón anterior
            
            
            pyautogui.click(query_x, query_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en New Query: {str(e)}")

    def click_gis_option(self):
        """Hacer clic en la opción GIS"""
        try:
            
            time.sleep(0.2)
            
            # Obtener posición de la ventana
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Usar la misma coordenada Y que New Query pero diferente X
            gis_x = window['x'] + int(size['width'] * 0.1)  # 10% del ancho
            gis_y = window['y'] + 160  # Misma altura que New Query
            
            
            pyautogui.click(gis_x, gis_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en GIS: {str(e)}")

    def click_address_option(self):
        """Hacer clic en la opción Address"""
        try:
            
            time.sleep(0.2)
            
            # Obtener posición de la ventana
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma coordenada X que New Query pero más abajo en Y
            address_x = window['x'] + int(size['width'] * 0.45)  # Igual que New Query
            address_y = window['y'] + 180  # 40px más abajo que New Query
            
            
            pyautogui.click(address_x, address_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Address: {str(e)}")

    def click_next_button(self):
        """Hacer clic en el botón Next"""
        try:
            
            time.sleep(0.7)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # El botón Next está en la parte inferior derecha
            next_x = window['x'] + int(size['width'] * 0.35)  # 75% del ancho
            next_y = window['y'] + 400  # 40px más abajo que Address
            
            
            pyautogui.click(next_x, next_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Next: {str(e)}")

    def click_actual_count(self):
        """Hacer clic en Actual Count"""
        try:
            print("Rellenando formulario para la busqueda de Dirección...")
            time.sleep(0.2)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma X que Address, pero más abajo en Y
            actual_x = window['x'] + int(size['width'] * 0.1)  # Igual que Address
            actual_y = window['y'] + 180  # Más abajo que el botón Next
            
            
            pyautogui.click(actual_x, actual_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Actual Count: {str(e)}")

    def click_name(self):
        """Hacer clic en el campo Name"""
        try:
            
            time.sleep(0.5)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma X que Actual Count, 100px más abajo
            name_x = window['x'] + int(size['width'] * 0.1)  # Igual que Actual Count
            name_y = window['y'] + 280  # 100px más abajo que Actual Count
            
            
            pyautogui.click(name_x, name_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Name: {str(e)}")

    def click_equals(self):
        """Hacer clic en el campo Equals"""
        try:
            
            time.sleep(0.2)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma Y que Actual Count, pero 20% en X
            equals_x = window['x'] + int(size['width'] * 0.2)  # 20% del ancho
            equals_y = window['y'] + 180  # Misma Y que Actual Count
            
            
            pyautogui.click(equals_x, equals_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Equals: {str(e)}")

    def click_is_like(self):
        """Hacer clic en Is Like"""
        try:
            
            time.sleep(0.2)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma X que Equals, 30px más en Y
            is_like_x = window['x'] + int(size['width'] * 0.2)  # Igual que Equals
            is_like_y = window['y'] + 220  # 40px más que Equals
            
            
            pyautogui.click(is_like_x, is_like_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Is Like: {str(e)}")

    def click_value(self):
        """Hacer clic en el campo Value y escribir la dirección"""
        try:
            print("Digitando Dirección a buscar...")
            time.sleep(0.2)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma Y que Equals, pero 30% en X
            value_x = window['x'] + int(size['width'] * 0.4)
            value_y = window['y'] + 180
            
            # Hacer clic y escribir la dirección
            pyautogui.click(value_x, value_y)
            time.sleep(0.2)
            
            # Asegurar que la dirección termine con *
            address_to_write = self.current_address if self.current_address.endswith('*') else f"{self.current_address}*"
            pyautogui.write(address_to_write)
            print(f"Dirección '{address_to_write}' escrita en Value")
            
        except Exception as e:
            print(f"Error escribiendo en Value: {str(e)}")

    def click_add_to_list(self):
        """Hacer clic en Add to List"""
        try:
            
            time.sleep(0.5)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma X que New Query, 20px más en Y que Value
            add_list_x = window['x'] + int(size['width'] * 0.45)  # Igual que New Query
            add_list_y = window['y'] + 200  # 20px más que Value
            
            
            pyautogui.click(add_list_x, add_list_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Add to List: {str(e)}")

    def click_search(self):
        """Hacer clic en Search"""
        try:
            print("Buscando Dirección...")
            time.sleep(0.3)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Misma X que New Query, pero en la parte inferior
            search_x = window['x'] + int(size['width'] * 0.45)  # Igual que New Query
            search_y = window['y'] + 380  # Parte inferior
            
            
            pyautogui.click(search_x, search_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Search: {str(e)}")

    def minimize_window(self):
        """Hacer clic en el botón minimizar"""
        try:
            print("Minimizando Formulario...")
            time.sleep(3)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Botón minimizar está al 50% del ancho y cerca de la parte superior
            minimize_x = window['x'] + int(size['width'] * 0.48)
            minimize_y = window['y'] + 100
            
            
            pyautogui.click(minimize_x, minimize_y)
            
            
        except Exception as e:
            print(f"Error minimizando ventana: {str(e)}")

    def click_go_to(self):
        """Hacer clic en el botón Go To (azul)"""
        try:
            print("Dirigiendonos a la Dirección...")
            time.sleep(1)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Botón Go To está en la parte superior
            goto_x = window['x'] + int(size['width'] * 0.28)  # 50% del ancho
            goto_y = window['y'] + 350  # Altura aproximada del botón azul
            
            
            pyautogui.click(goto_x, goto_y)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Go To: {str(e)}")

    def set_scale(self):
        """Modificar la escala a 1:188 y presionar Enter"""
        try:
            print("Modificando escala...")
            time.sleep(1)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Coordenadas para el campo de escala
            scale_x = window['x'] + int(size['width'] * 0.35)  # Posición del campo de escala
            scale_y = window['y'] + 130  # Altura del campo de escala
            
            # Hacer clic y escribir 88 (se agregará al 1:1 existente)
            
            pyautogui.click(scale_x, scale_y)
            time.sleep(0.5)
            pyautogui.write("88")
            time.sleep(0.5)
            pyautogui.press('enter')  # Presionar Enter después de escribir
            
            
        except Exception as e:
            print(f"Error modificando escala: {str(e)}")

    def draw_selection_box(self):
        """Hacer doble clic en el centro donde estaría el círculo verde"""
        try:
            print("Seleccionando dirección")
            time.sleep(1.5)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Calcular el centro del área
            center_x = window['x'] + int(size['width'] * 0.327)
            center_y = window['y'] + int(size['height'] * 0.427)
            
            # Primero mover el mouse a la posición
            pyautogui.moveTo(center_x, center_y)
            time.sleep(1)  # Esperar que el mouse esté estable
            
            # Realizar los dos clics con intervalo
            
            pyautogui.click()
            time.sleep(0.7)  # Esperar entre clics
            
            pyautogui.click()
            
            
            
        except Exception as e:
            print(f"Error haciendo clics en el centro: {str(e)}")

    def get_association_value(self, save_to_excel=False):
        """Hacer clic en el ícono de Word y Export to Excel"""
        try:
            
            time.sleep(2)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # El icono está a un 60% del ancho y alineado con el botón Go To
            word_x = window['x'] + int(size['width'] * 0.49)
            word_y = window['y'] + 350
            
            
            pyautogui.click(word_x, word_y)
            
            time.sleep(1)
            
            # Clic en Export to Excel (misma X que Word, 15px más abajo)
            export_x = word_x
            export_y = word_y + 15
            
            
            pyautogui.click(export_x, export_y)
            time.sleep(0.5)
            
            # Clic en Save File (misma X pero 80px menos)
            save_y = export_y - 80  # Modificado a -80px
            
            pyautogui.click(export_x, save_y)
            time.sleep(0.5)
            
            
        except Exception as e:
            print(f"Error haciendo clic en Word/Export: {str(e)}")

    def handle_download_dialog(self):
        """Manejar el diálogo de descarga"""
        try:
            print("Esperando diálogo de descarga...")
            time.sleep(1)
            # Solo presionar enter ya que Save File está seleccionado
            pyautogui.press('enter')
            print("Descarga iniciada")
            
        except Exception as e:
            print(f"Error manejando diálogo de descarga: {str(e)}")

    def handle_excel_file(self, timeout=10):  # Changed from 30 to 10 seconds
        """Detectar y mover el archivo Excel data(#).xls más reciente"""
        try:
            print("Esperando archivo Excel 'data.xls'...")
            time.sleep(3)
            start_time = datetime.now()
            
            # Crear carpeta de resultados si no existe
            os.makedirs(self.results_folder, exist_ok=True)
            

            
            while (datetime.now() - start_time).seconds < timeout:
                # Buscar archivos data.xls
                excel_files = glob.glob(os.path.join(self.download_folder, 'data*.xls'))
                
                if excel_files:
                    print(f"Archivos encontrados: {excel_files}")
                    
                    # Función para extraer el número del paréntesis
                    def get_file_number(filename):
                        import re
                        match = re.search(r'data\((\d+)\)\.xls$', filename)
                        return int(match.group(1)) if match else 0
                    
                    # Ordenar archivos por número de versión, los sin número serán 0
                    latest_file = max(excel_files, key=lambda f: (
                        get_file_number(f),  # Primero por número
                        os.path.getmtime(f)  # Luego por fecha de modificación
                    ))
                    
                    print(f"Archivo más reciente encontrado: {latest_file}")
                    time.sleep(1)
                    
                    try:
                        new_path = os.path.join(self.results_folder, 'data.xls')
                        shutil.move(latest_file, new_path)
                        print(f"Archivo movido exitosamente a: {new_path}")
                        return new_path
                    except PermissionError:
                        print("Archivo aún en uso, esperando...")
                        time.sleep(1)
                        continue
                    except Exception as move_error:
                        print(f"Error moviendo archivo: {str(move_error)}")
                        return None
                
                time.sleep(0.5)
            
            print(f"Timeout después de {timeout} segundos")
            # Si no se encontró el archivo, agregar la dirección actual a not_found_addresses
            if self.current_address not in self.not_found_addresses:
                self.not_found_addresses.append(self.current_address)
            return None
            
        except Exception as e:
            print(f"Error manejando archivo Excel: {str(e)}")
            # También agregar a not_found_addresses en caso de error
            if self.current_address not in self.not_found_addresses:
                self.not_found_addresses.append(self.current_address)
            return None

    def get_nap_status(self):
        """Consultar estado de la NAP al usuario"""
        print("\nEstado de la NAP:")
        print("1. Activa")
        print("2. En Diseño")
        print("3. Activa y En Diseño")
        print("4. Otro")
        
        while True:
            try:
                opcion = input("\nSeleccione una opción (1-4): ")
                if opcion not in ['1', '2', '3', '4']:
                    print("Opción inválida. Por favor seleccione 1, 2, 3 o 4.")
                    continue
                
                if opcion == '1':
                    return 'Activa'
                elif opcion == '2':
                    return 'En Diseño'
                elif opcion == '3':
                    return 'Activa y En Diseño'
                else:
                    estado_custom = input("Ingrese el estado personalizado: ")
                    return estado_custom
                    
            except Exception as e:
                print(f"Error: {str(e)}")
                print("Por favor intente nuevamente.")

    def process_excel_files(self, data_path):
        """Procesar archivos Excel y almacenar información"""
        try:
            print("Procesando archivos Excel...")
            
            try:
                workbook = xlrd.open_workbook(data_path)
                sheet = workbook.sheet_by_name('Address')
                
                # Buscar índices de las columnas en hoja Address
                header_row = sheet.row_values(0)
                name_col = header_row.index('Name')
                asociacion_col = header_row.index('Asociacion')
                
                # Buscar la dirección actual y obtener su valor de Asociacion
                asociacion_valor = None
                for row in range(1, sheet.nrows):
                    name_value = str(sheet.cell_value(row, name_col))
                    if self.current_address.strip('*') in name_value:
                        asociacion_valor = sheet.cell_value(row, asociacion_col)
                        break

                # Determinar valor de asociación
                if asociacion_valor is None:
                    valor_final = 'No'
                else:
                    valor_final = 'Si' if str(asociacion_valor).lower() == 'true' else 'No'

                # Verificar hoja Molecula y extraer datos
                tiene_hoja_molecula = 'Molecula' in workbook.sheet_names()
                codigo_macro = propietario_valor = estado_nap = 'N/A'
                
                if tiene_hoja_molecula:
                    sheet_molecula = workbook.sheet_by_name('Molecula')
                    header_molecula = sheet_molecula.row_values(0)
                    type_col = header_molecula.index('Type')
                    codigo_col = header_molecula.index('Codigo')
                    propietario_col = header_molecula.index('Propietario')
                    
                    # Buscar fila con Type = 'Macro Cell'
                    for row in range(1, sheet_molecula.nrows):
                        if str(sheet_molecula.cell_value(row, type_col)) == 'Macro Cell':
                            codigo_macro = sheet_molecula.cell_value(row, codigo_col)
                            propietario_valor = sheet_molecula.cell_value(row, propietario_col)
                            break
                    
                    # Solo preguntar estado NAP si hay molécula
                    estado_nap = self.get_nap_status()
                
                # Almacenar datos
                self.address_data[self.current_address] = {
                    'dentro_molecula': 'Si' if tiene_hoja_molecula else 'No',
                    'asociado': valor_final,
                    'propietario': propietario_valor,
                    'codigo_macro': codigo_macro,
                    'estado_nap': estado_nap
                }
                
                print(f"Datos almacenados para dirección {self.current_address}")
                
            except Exception as e:
                print(f"Error procesando Excel: {str(e)}")
                
        except Exception as e:
            print(f"Error general: {str(e)}")

    def reduce_address_parameters(self, address):
        """Reducir parámetros de la dirección"""
        parts = address.split()
        if len(parts) <= 3:  # Si ya tiene 3 o menos parámetros, no se puede reducir más
            return None
        return ' '.join(parts[:-1])  # Eliminar último parámetro

    def process_search_result(self, address):
        """Procesar resultado de búsqueda y manejar reducciones"""
        # Validar número de parámetros primero
        if not self.validate_address_parameters(address):
            print(f"Dirección {address} tiene menos de 4 parámetros - marcando como no encontrada")
            self.not_found_addresses.append(address)
            return False

        original_address = address
        current_try = address
        
        while True:
            self.current_address = current_try
            print(f"\nBuscando dirección: {current_try}")
            
            # Realizar búsqueda
            self.click_actual_count()
            time.sleep(1)
            self.click_name()
            time.sleep(1)
            self.click_equals()
            time.sleep(1)
            self.click_is_like()
            time.sleep(1)
            self.click_value()
            time.sleep(1)
            self.click_add_to_list()
            self.click_search()
            self.minimize_window()
            self.click_go_to()
            self.set_scale()
            self.draw_selection_box()
            
            # Descargar y procesar Excel
            self.get_association_value(save_to_excel=True)
            self.handle_download_dialog()
            excel_path = self.handle_excel_file()
            
            if excel_path and os.path.exists(excel_path):
                print(f"Información encontrada para: {current_try}")
                self.found_addresses[original_address] = current_try
                return True
                
            # Si no se encontró, reducir parámetros
            next_try = self.reduce_address_parameters(current_try)
            if not next_try:
                print(f"No se encontró información para: {original_address}")
                self.not_found_addresses.append(original_address)
                return False
                
            current_try = next_try
            print(f"Intentando con menos parámetros: {current_try}")

    def update_backlog(self):
        """Actualizar Backlog con toda la información recolectada"""
        try:
            print("\nActualizando Backlog con toda la información...")
            backlog_path = 'Backlog250525-filtradocobertura.xlsx'
            
            workbook = openpyxl.load_workbook(backlog_path)
            sheet = workbook['Detalle']
            
            # Definir columnas
            param_col = 26      # Columna Z - Parametrización
            dentro_mol_col = 27 # Columna AA - Dentro molecula Gpon
            asoc_col = 28       # Columna AB - Asociado molecula
            prop_mol_col = 29   # Columna AC - Propietario Molecula
            nombre_mol_col = 30 # Columna AD - Nombre de la moleula
            estado_mol_col = 31 # Columna AE - Estado Molecula
            
            red_font = openpyxl.styles.Font(color='FF0000')
            
            # Actualizar datos para cada dirección incluyendo duplicados
            for direccion, indices in self.address_indices.items():
                data = None
                if direccion in self.address_data:
                    data = self.address_data[direccion]
                elif direccion in self.not_found_addresses:
                    # Marcar todas las instancias como no encontradas
                    for idx in indices:
                        row = idx + 2  # +2 porque Excel empieza en 1 y tiene header
                        cell = sheet.cell(row=row, column=param_col)
                        cell.font = red_font
                        cell.value = f"{direccion} no aparece direccion"
                        for col in [dentro_mol_col, asoc_col, prop_mol_col, nombre_mol_col, estado_mol_col]:
                            sheet.cell(row=row, column=col, value='')
                    continue
                
                # Actualizar todas las instancias con la misma información
                if data:
                    for idx in indices:
                        row = idx + 2
                        sheet.cell(row=row, column=dentro_mol_col, value=data['dentro_molecula'])
                        sheet.cell(row=row, column=asoc_col, value=data['asociado'])
                        sheet.cell(row=row, column=prop_mol_col, value=data['propietario'])
                        sheet.cell(row=row, column=nombre_mol_col, value=data['codigo_macro'])
                        sheet.cell(row=row, column=estado_mol_col, value=data['estado_nap'])
                    print(f"Actualizada dirección {direccion} en {len(indices)} ubicaciones")
            
            workbook.save(backlog_path)
            print("Backlog actualizado exitosamente")
            
        except Exception as e:
            print(f"Error actualizando Backlog: {str(e)}")

    def update_backlog_batch(self, batch_addresses):
        """Actualizar Backlog para un lote de direcciones"""
        try:
            print("\nActualizando Backlog para el lote actual...")
            backlog_path = 'Backlog_GPON_FILTRADO.xlsx'
            
            if not os.path.exists(backlog_path):
                print(f"\n❌ Error: No se encuentra el archivo '{backlog_path}'")
                print("Por favor, verifique que el archivo esté en la carpeta correcta.")
                print(f"Carpeta actual: {os.getcwd()}")
                return
            
            workbook = openpyxl.load_workbook(backlog_path)
            sheet = workbook['GPON']  # Cambiar a hoja GPON
            
            # Definir columnas con nombres correctos
            param_col = None  # Se buscará en el archivo
            dentro_mol_col = None
            asoc_col = None
            prop_mol_col = None
            nombre_mol_col = None
            estado_mol_col = None
            
            # Encontrar índices de columnas
            for idx, cell in enumerate(sheet[1], 1):
                col_name = str(cell.value).strip() if cell.value else ''
                if col_name == 'Prametrización':
                    param_col = idx
                elif col_name == 'Dentro molecula Gpon':
                    dentro_mol_col = idx
                elif col_name == 'Asociado molecula':
                    asoc_col = idx
                elif col_name == 'Propietario Molecula':
                    prop_mol_col = idx
                elif col_name == 'Nombre de la moleula':
                    nombre_mol_col = idx
                elif col_name == 'Estado Molecula':
                    estado_mol_col = idx
            
            red_font = openpyxl.styles.Font(color='FF0000')
            
            # Actualizar datos
            for direccion in batch_addresses:
                if direccion not in self.address_indices:
                    continue
                    
                indices = self.address_indices[direccion]
                data = None
                
                if direccion in self.address_data:
                    data = self.address_data[direccion]
                elif direccion in self.not_found_addresses:
                    # Marcar como no encontrada
                    for idx in indices:
                        row = idx + 2
                        cell = sheet.cell(row=row, column=param_col)
                        cell.font = red_font
                        cell.value = f"{direccion} no aparece direccion"
                        sheet.cell(row=row, column=dentro_mol_col, value='N/A')
                        sheet.cell(row=row, column=asoc_col, value='N/A')
                        sheet.cell(row=row, column=prop_mol_col, value='N/A')
                        sheet.cell(row=row, column=nombre_mol_col, value='N/A')
                        sheet.cell(row=row, column=estado_mol_col, value='N/A')
                    continue
                
                if data:
                    for idx in indices:
                        row = idx + 2
                        sheet.cell(row=row, column=dentro_mol_col, value=data['dentro_molecula'])
                        sheet.cell(row=row, column=asoc_col, value=data['asociado'])
                        sheet.cell(row=row, column=prop_mol_col, value=data['propietario'])
                        sheet.cell(row=row, column=nombre_mol_col, value=data['codigo_macro'])
                        sheet.cell(row=row, column=estado_mol_col, value=data['estado_nap'])
            
            workbook.save(backlog_path)
            print("✅ Excel actualizado exitosamente")
            
        except Exception as e:
            print(f"Error actualizando Backlog por lotes: {str(e)}")

    def validate_backlog_file(self, filename: str) -> bool:
        """Validar existencia del archivo Backlog"""
        if not os.path.exists(filename):
            print(f"\n❌ Error: No se encuentra el archivo '{filename}'")
            print("Por favor, verifique que el archivo esté en la carpeta correcta.")
            print(f"Carpeta actual: {os.getcwd()}")
            return False
        return True

    def setup_driver(self):
        try:
            # Validar existencia del archivo Backlog antes de empezar
            backlog_filename = "Backlog_GPON_FILTRADO.xlsx"
            if not self.validate_backlog_file(backlog_filename):
                return

            print("Iniciando Firefox...")
            profile = webdriver.FirefoxProfile()
            profile.set_preference("browser.startup.page", 0)
            
            # Usar sintaxis de Selenium 2.53.6
            self.driver = webdriver.Firefox(firefox_profile=profile)
            print("Firefox iniciado")
            
            # Navegar y hacer login
            self.driver.get(self.url)
            self.login()
            
            # Activar Silverlight
            self.handle_silverlight()
            
            # Hacer clic en el botón default
            self.click_default_button()
            
            # Hacer clic en New Query
            self.click_new_query()
            
            # Hacer clic en GIS
            self.click_gis_option()
            
            # Hacer clic en Address
            self.click_address_option()
            
            # Hacer clic en Next
            self.click_next_button()
            
            # Calcular tamaño de lotes para actualización del Excel
            total_addresses = len(self.addresses)
            batch_size = max(1, total_addresses // 5)  # Dividir en 5 partes
            print(f"\nℹ️ Total de direcciones a procesar: {total_addresses}")
            print(f"ℹ️ Se actualizará el Excel cada {batch_size} direcciones procesadas")
            
            current_batch = []
            
            for idx, address in enumerate(self.addresses, 1):
                self.current_address = address
                print(f"\nProcesando dirección {idx}/{len(self.addresses)}: {address}")
                
                # Validar número de parámetros
                if not self.validate_address_parameters(address):
                    print(f"Dirección {address} tiene menos de 4 parámetros - marcando como no encontrada")
                    self.not_found_addresses.append(address)
                    self.address_data[address] = {
                        'dentro_molecula': 'N/A',
                        'asociado': 'N/A',
                        'propietario': 'N/A',
                        'codigo_macro': 'N/A',
                        'estado_nap': 'N/A'
                    }
                    current_batch.append(address)
                else:
                    # Procesar dirección normalmente
                    self.click_actual_count()
                    time.sleep(1)
                    self.click_name()
                    time.sleep(1)
                    self.click_equals()
                    time.sleep(1)
                    self.click_is_like()
                    time.sleep(1)
                    self.click_value()
                    time.sleep(1)
                    self.click_add_to_list()
                    self.click_search()
                    self.minimize_window()
                    self.click_go_to()
                    self.set_scale()
                    self.draw_selection_box()
                    
                    # Descargar y procesar Excel
                    self.get_association_value(save_to_excel=True)
                    self.handle_download_dialog()
                    excel_path = self.handle_excel_file()
                    if excel_path:
                        self.process_excel_files(excel_path)
                        current_batch.append(address)
                    else:
                        print(f"No se encontró información para: {address}")
                        self.address_data[address] = {
                            'dentro_molecula': 'N/A',
                            'asociado': 'N/A',
                            'propietario': 'N/A',
                            'codigo_macro': 'N/A',
                            'estado_nap': 'N/A'
                        }
                        self.not_found_addresses.append(address)
                        current_batch.append(address)
                
                # Actualizar Excel cada batch_size direcciones o al final
                if len(current_batch) >= batch_size or idx == len(self.addresses):
                    print(f"\n📝 Actualizando Excel (procesadas {len(current_batch)} direcciones)...")
                    try:
                        self.update_backlog_batch(current_batch)
                        print("✅ Excel actualizado exitosamente")
                        current_batch = []  # Limpiar batch actual
                    except FileNotFoundError:
                        print(f"\n❌ Error: No se encuentra el archivo '{backlog_filename}'")
                        print("Por favor, verifique que el archivo esté en la carpeta correcta.")
                        print(f"Carpeta actual: {os.getcwd()}")
                        if input("\n¿Desea continuar con la automatización? (s/n): ").lower() != 's':
                            break
                    except Exception as e:
                        print(f"\n❌ Error actualizando Excel: {str(e)}")
                        if input("\n¿Desea continuar con la automatización? (s/n): ").lower() != 's':
                            break
                
                # Restaurar para siguiente dirección si no es la última
                if idx < len(self.addresses):
                    self.click_default_button()
                    time.sleep(1)
                    self.click_clear()
            
            input("Presiona Enter para cerrar...")
            
        except Exception as e:
            print(f"Error: {str(e)}")
            raise
        finally:
            if hasattr(self, 'driver'):
                self.driver.quit()

    def validate_address_parameters(self, address: str) -> bool:
        """Validar que la dirección tenga al menos 4 parámetros"""
        if not address:
            return False
        # Dividir por espacios y filtrar elementos vacíos
        params = [p for p in address.split() if p.strip()]
        return len(params) >= 4

    def click_clear(self):
        """Hacer clic en el botón Clear"""
        try:
            print("Haciendo clic en Clear...")
            time.sleep(0.5)
            
            window = self.driver.get_window_position()
            size = self.driver.get_window_size()
            
            # Coordenadas para el botón Clear
            clear_x = window['x'] + int(size['width'] * 0.45)
            clear_y = window['y'] + 300  # Más abajo que Add to List
            
            print(f"Haciendo clic en Clear: x={clear_x}, y={clear_y}")
            pyautogui.click(clear_x, clear_y)
            print("Clic realizado en Clear")
            time.sleep(0.5)  # Esperar que se complete el clear
            
        except Exception as e:
            print(f"Error haciendo clic en Clear: {str(e)}")

def run_automation():
    try:
        automation = AddressAutomation()
        return automation
    except Exception as e:
        print(f"❌ Error en la automatización: {str(e)}")
        return None

if __name__ == "__main__":
    run_automation()